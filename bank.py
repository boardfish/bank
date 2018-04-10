from pymonzo import MonzoAPI  # Import Monzo Class
from dateutil import parser
import pytz
import config as cfg
import csv
import datetime
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill
from openpyxl import Workbook
from openpyxl import load_workbook
from dateutil import parser


# Retrieves the list of transactions from Monzo, optionally with merchants.
def init_monzo(merchants=False):
    client = MonzoAPI()
    # Get the ID of the first account linked to the access token
    account_id = client.accounts()[1].id
    print("[Monzo] Retrieving transactions...")
    transactions = client.transactions(
        account_id)  # Get client.transaction(
    print("[Monzo] Retrieving merchants...")
    for item in transactions:
        identifier = item.merchant
        if merchants:
            item.merchant = client.transaction(
                item.id, expand_merchant=False).merchant
            identifier = item.merchant.name
        print("[Monzo] Retrieved merchant {}.".format(identifier))
    print("Done.")
    return transactions


def parse_monzo(transactions):
    transactionsParsed = []
    for item in transactions:
        # Merchant
        try:
            merchant = item.merchant.name
        except AttributeError:
            merchant = "No name"
        except TypeError:
            merchant = "No merchant for this item"
        # Date
        date = item.created
        amount = int(item.amount)
        transactionsParsed.append({
            'date': date,
            'transaction': amount,
            'merchant': merchant})
    return transactionsParsed


def sort_chronologically(transactions):
    # transactions.sort(key=lambda item:item['date'])
    transactions = sorted(transactions, key=lambda k: k['date'])
    return transactions


def sort_months(transactions):
    # Store the dates in stored_transactions with Y/M dates
    sorted_transactions = {}
    for transaction in transactions:
        # Make a list of months
        month = transaction['date'].strftime("%y/%m")
        if month in sorted_transactions:
            sorted_transactions[month].append(transaction)
        else:
            sorted_transactions[month] = [transaction]
    print("sorted")
    ordered_transactions = {}
    for key in sorted(sorted_transactions.keys()):
        print(key)
        ordered_transactions[key] = sorted_transactions[key]
    print("sorted")
    print(ordered_transactions)
    return ordered_transactions

# Formatting and displaying


def to_2sf(value):
    return float(str.format('{0:.2f}', value/100))


def to_pounds(pence):
    value = to_2sf(pence)
    if str(value[0]) == '-':
        value = value[1:]
        prefix = "-£"
    else:
        prefix = "+£"
    return prefix + value


def format_for_display(transactions):
    for row in transactions:
        row['date'] = row['date'].strftime('%d/%m/%y')
        row['transaction'] = to_pounds(row['transaction'])
        print(row)
    return transactions

# Exporting


def write_to_csv(transactions, filename):
    with open(filename, 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=';',
                               quotechar='|', quoting=csv.QUOTE_MINIMAL)
        csvwriter.writerow(['Date', 'Merchant', 'Transaction', 'Balance'])
        cumulative_total = 0
        for row in transactions:
            cumulative_total += int(row['transaction'])
            csvwriter.writerow(
                [str(row['date']), row['merchant'], row['transaction'], cumulative_total])


def beautify(transactions):
    col_width = 40
    for row in sort_chronologically(transactions):
        print("".join(str(row[key]).ljust(col_width) for key in row))

# Exporting: Excel


def excel_autofit(ws):
    # from https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length and str(cell.value)[0] != '=':
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def excel_summary_headers(ws, col, transactions):
    row_iterator = 1
    # for key in transactions:
    #    print("Key: {}".format(key))
    #     header_row.append(key)
    # ws.append(header_row)
    ws.cell(row=row_iterator, column=1, value="Category")
    row_iterator += 1
    # Row titles
    for category in cfg.outgoing_categories:
        ws.cell(row=row_iterator, column=1, value=category)
        row_iterator += 1
    ws.cell(row=row_iterator, column=1, value='Total Outgoings')
    row_iterator += 1
    for category in cfg.income_categories:
        ws.cell(row=row_iterator, column=1, value=category)
        row_iterator += 1
    ws.cell(row=row_iterator, column=1, value='Total Income')
    row_iterator += 1
    ws.cell(row=row_iterator, column=1, value='Balance')
    row_iterator += 1


def excel_summary_totals(ws, col, columns):
    row_iterator = 2
    ws.cell(row=row_iterator, column=1, value='Total')
    # Row titles
    summary = [
        'B',
        chr(columns + ord('A'))
    ]
    for x in range(0, len(cfg.outgoing_categories)+len(cfg.income_categories)+3):
        val = '=SUM({1}{0}:{2}{0})'.format(
            row_iterator, summary[0], summary[1])
        ws.cell(row=row_iterator, column=col, value=val)
        row_iterator += 1
    val = '=SUM({1}2:{1}{0})'.format(row_iterator-1, chr(ord('A') - 1 + col))


def excel_summary_column(transactions, ws, source_sheet, col):
    row_iterator = 2  # leave header row
    # should return e.g. 'B' for input of 2
    col_letter = chr(col + ord('A') - 1)
    # append sum of outgoings
    outgoings = [row_iterator, 0]
    for category in cfg.outgoing_categories:
        val = '=SUMPRODUCT(({2}!D2:D{1}="{0}")*{2}!C2:C{1})'.format(
            category, len(transactions)+1, source_sheet)
        ws.cell(row=row_iterator, column=col, value=val)
        row_iterator += 1
    outgoings[1] = row_iterator - 1
    val = '=SUM({0}{1}:{0}{2})'.format(col_letter, outgoings[0], outgoings[1])
    ws.cell(row=row_iterator, column=col, value=val)
    row_iterator += 1
    incomes = [row_iterator, 0]
    for category in cfg.income_categories:
        val = '=SUMPRODUCT(({2}!D2:D{1}="{0}")*{2}!C2:C{1})'.format(
            category, len(transactions)+1, source_sheet)
        ws.cell(row=row_iterator, column=col, value=val)
        row_iterator += 1
    incomes[1] = row_iterator - 1
    # append sum of incomes
    val = '=SUM({0}{1}:{0}{2})'.format(col_letter, incomes[0], incomes[1])
    ws.cell(row=row_iterator, column=col, value=val)
    row_iterator += 1
    # append sum of both
    prev_col = chr(ord(col_letter)-1)
    if prev_col == "A":
        val = '={0}{1}+{0}{2}'.format(col_letter, outgoings[1]+1, incomes[1])
    else:
        val = '={3}{4}+{0}{1}+{0}{2}'.format(col_letter,
                                             outgoings[1]+1, incomes[1], prev_col, row_iterator)
    excel_format_currency(ws.cell(row=row_iterator, column=col, value=val))
    # Formatting for income and balance totals
    for row in ws.iter_rows(min_row=incomes[0], max_row=incomes[1]):
        for cell in row:
            cell.style = "20 % - Accent1"
            excel_format_currency(cell)
    for row in ws.iter_rows(min_row=incomes[1]+1, max_row=incomes[1]+1):
        for cell in row:
            cell.style = "60 % - Accent1"
            excel_format_currency(cell)
    for row in ws.iter_rows(min_row=incomes[1]+2, max_row=incomes[1]+2):
        for cell in row:
            cell.style = "60 % - Accent3"
            excel_format_currency(cell)
    # Header row of spending summary
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.style = "60 % - Accent4"
            excel_format_currency(cell)
    # Outgoing totals
    for row in ws.iter_rows(min_row=2, max_row=len(cfg.outgoing_categories)+1):
        for cell in row:
            cell.style = "20 % - Accent2"
            excel_format_currency(cell)
    for row in ws.iter_rows(min_row=len(cfg.outgoing_categories)+2, max_row=len(cfg.outgoing_categories)+2):
        for cell in row:
            cell.style = "60 % - Accent2"
            excel_format_currency(cell)
    excel_autofit(ws)


def excel_format_currency(cell):
    cell.number_format = '£#,##0.00'


def excel_export(transactions, filename):
    # Formatting cells
    # Basic styles
    header_font = Font(bold=True)
    # init
    wb = Workbook()
    # grab the active worksheet
    # Spending Summary
    ws = wb.active
    ws.title = "Spending Summary"
    excel_summary_headers(ws, 1, transactions)
    summary_column = 2
    sorted_transactions = sort_months(transactions)
    for month, monthStatement in sorted_transactions.items():
        newkey = datetime.datetime.strptime(month, "%y/%m").strftime("%B%y")
        excel_summary_column(monthStatement, ws, newkey, summary_column)
        summary_column += 1
    # Transaction List
        ws1 = wb.create_sheet(newkey)
        ws1.append(['Date', 'Merchant', 'Transaction', "Category"])
        for transaction in monthStatement:
            ws1.append([transaction['date'], transaction['merchant'],
                        to_2sf(transaction['transaction'])])
        for row in ws1.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                excel_format_currency(cell)
        # Coloring header row of transaction list
        for row in ws1.iter_rows(min_row=1, max_col=4, max_row=1):
            for cell in row:
                cell.font = header_font
                cell.style = "60 % - Accent1"
        excel_autofit(ws1)
    print("summary_column:", summary_column)
    excel_summary_totals(ws, summary_column, len(sorted_transactions.keys()))
    # Save the file
    wb.save(filename)
    print("Saved. Don't forget to check cell references - it might not be perfect.")
    import subprocess
    subprocess.Popen(["libreoffice", filename])


# INIT
print("Fetching data from Monzo...")
t = init_monzo()
print("Exporting to Excel...")
excel_export(parse_monzo(t), "monzo.xlsx")
# monzoTransactions = parse_monzo(init_monzo())
# santanderTransactions = init_santander(cfg.santander_statement)
# transactions = santanderTransactions + monzoTransactions
# PRINT
# excel_export(sort_months(transactions), 'sample.xlsx')
